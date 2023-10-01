from openpyxl import load_workbook

wb = load_workbook(filename = "template.xlsx")

ws = wb.active

print(tuple(ws.rows))
ws['B5'] = "Comet Task Force"

wb.save("output.xlsx")