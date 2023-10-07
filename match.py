import numpy as np
import argparse
import glob
import cv2

from img_to_stat import ImgToStat
from team_detect import TeamDetector
from openpyxl import load_workbook

# construct the argument parser and parse the arguments
ap = argparse.ArgumentParser()
ap.add_argument("-i", "--images", required=True, help="Path to images to be read")
ap.add_argument("-d", "--debug", required=False, help="Enables debug", action='store_true')
args = vars(ap.parse_args())

def sortByY(e):
	return e[1]

# load the image image, convert it to grayscale, and detect edges
template = cv2.imread("templates/victory.PNG")
template = cv2.cvtColor(template, cv2.COLOR_BGR2GRAY)
template = cv2.Canny(template, 50, 200)
(tH, tW) = template.shape[:2]
#cv2.imshow("Template", template)

# intialize team detector
td = TeamDetector('rosters.csv')

wb = load_workbook(filename = "template.xlsx")

ws = wb.active
active_row = 2

# loop over the images to find the template in
for imagePath in glob.glob(args["images"] + "/*.jpg"):
	image = cv2.imread(imagePath)
	stats = ImgToStat(image, cv2.imread("templates/victory.PNG"), td)
	# load data onto sheet
	for i, data in enumerate(stats.get_data()):
		target_cell = ws.cell(row=active_row, column=i+1, value=data)
	active_row += 1
	wb.save(args["images"] + "/output.xlsx")

for imagePath in glob.glob(args["images"] + "/*.png"):
	image = cv2.imread(imagePath)
	print(imagePath)
	stats = ImgToStat(image, cv2.imread("templates/victory.PNG"), td)
	# load data onto sheet
	for i, data in enumerate(stats.get_data()):
		target_cell = ws.cell(row=active_row, column=i+1, value=data)
	active_row += 1
	wb.save(args["images"] + "/output.xlsx")

print("Output file to: " + args["images"] + "/output.xlsx")
